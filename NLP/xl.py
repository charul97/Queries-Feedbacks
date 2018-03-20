import os, json
import pandas as pd
from pprint import pprint
import xlrd
from xlrd import open_workbook
from openpyxl import worksheet
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
import nltk.stem.snowball
import string
import openpyxl

workbook = openpyxl.load_workbook('/home/adi-sin/Desktop/dummy.xlsx')
wb=workbook.get_sheet_by_name('dummy')
r = wb.max_row
dict ={}
for i in range(1,r+1):
	ans= wb.cell(row= i, column=2).value
	ques= wb.cell(row=i, column=1).value
	dict[ques]=ans




stopwords = nltk.corpus.stopwords.words('english')
stopwords.extend(string.punctuation)
stopwords.append('')

qlist = []

answer = []

output_file=open('/home/adi-sin/Desktop/Namastey/data2', 'r')
for items in output_file:
	sent = sent_tokenize(items)
	qlist.append(sent)

size = len(qlist)	

for i in range(1,size):

	new_dict = {}
	x=[]

	for key,value in dict.items():
		a=key
		b = qlist[i]


		tokenizera = word_tokenize(a)
		tokenizerb = word_tokenize(str(b))
  
		tokens_a = [token.lower().strip(string.punctuation) for token in tokenizera \
                    if token.lower().strip(string.punctuation) not in stopwords]
		tokens_b = [token.lower().strip(string.punctuation) for token in tokenizerb \
                    if token.lower().strip(string.punctuation) not in stopwords]

        # Calculate Jaccard similarity
		ratio = len(set(tokens_a).intersection(tokens_b)) / float(len(set(tokens_a).union(tokens_b)))
		x.append(ratio)
		new_dict[ratio] = value
	

	y=max(x,key=float)
	
	
    
	for key,value in new_dict.items():
		if(y > 0.5):
			if(key == y):
				print("Question:")
				print(qlist[i])
				print()
				print(key)
				print("Answer")
				print(value)
				answer.append(value)
				


print(str(answer))

with open('/home/adi-sin/Desktop/Namastey/query_reply', 'w') as ans:
	ans.write(str(answer))			

